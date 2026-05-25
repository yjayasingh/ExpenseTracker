"""Export expenses to Excel (.xlsx)."""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import database as db

HEADERS = [
    "ID",
    "Date",
    "Description",
    "Category",
    "Amount (LKR)",
    "Receipt",
    "Created At",
]


def build_workbook(expenses):
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for e in expenses:
        ws.append([
            e["id"],
            e["expense_date"],
            e["description"],
            e["category"],
            e["amount"],
            "Yes" if e.get("receipt_path") else "No",
            e.get("created_at", ""),
        ])

    total_row = len(expenses) + 2
    ws.cell(row=total_row, column=4, value="Total")
    ws.cell(row=total_row, column=4).font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=sum(e["amount"] for e in expenses))
    ws.cell(row=total_row, column=5).font = Font(bold=True)

    for col in range(1, len(HEADERS) + 1):
        letter = get_column_letter(col)
        max_len = len(HEADERS[col - 1])
        for row in ws.iter_rows(min_col=col, max_col=col, min_row=1):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 50)

    return wb


def export_all_expenses():
    expenses = db.get_expenses()
    wb = build_workbook(expenses)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"expenses_{date.today().isoformat()}.xlsx"
    return buffer, filename
