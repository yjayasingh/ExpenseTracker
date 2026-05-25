"""Tests for export_excel module."""

from io import BytesIO

from openpyxl import load_workbook

import database as db
import export_excel


def test_build_workbook_headers_and_rows(temp_db):
    db.add_expense(100, "Coffee", "Food", "2026-05-01")
    db.add_expense(250, "Bus", "Transport", "2026-05-02", "r.png")

    expenses = db.get_expenses()
    wb = export_excel.build_workbook(expenses)
    ws = wb.active

    assert ws.title == "Expenses"
    assert [c.value for c in ws[1]] == export_excel.HEADERS
    # Expenses ordered by date descending (newest first)
    assert ws[2][1].value == "2026-05-02"
    assert ws[2][2].value == "Bus"
    assert ws[2][5].value == "Yes"
    assert ws[3][1].value == "2026-05-01"
    assert ws[3][2].value == "Coffee"

    total_row = len(expenses) + 2
    assert ws.cell(row=total_row, column=4).value == "Total"
    assert ws.cell(row=total_row, column=5).value == 350


def test_build_workbook_empty(temp_db):
    wb = export_excel.build_workbook([])
    ws = wb.active
    assert ws[1][0].value == "ID"
    assert ws.cell(row=2, column=4).value == "Total"
    assert ws.cell(row=2, column=5).value == 0


def test_export_all_expenses(temp_db):
    db.add_expense(500, "Test", "Other", "2026-05-25")

    buffer, filename = export_excel.export_all_expenses()

    assert filename.startswith("expenses_")
    assert filename.endswith(".xlsx")
    assert len(buffer.getvalue()) > 0

    wb = load_workbook(BytesIO(buffer.getvalue()))
    assert wb.active.max_row >= 2
